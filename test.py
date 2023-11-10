from openpyxl import Workbook, load_workbook


print('start')
filename = "rez.xlsx"
wb = load_workbook(filename=filename)
sheet = wb.active
print(sheet.cell(row=1, column=1).value)
i = 2
k = 0
w = 0
m = 0
p = 0
r = 0
while sheet.cell(row=i, column=1).value:
    a = max(
        float(sheet.cell(row=i, column=1).value),
        float(sheet.cell(row=i, column=2).value),
        float(sheet.cell(row=i, column=3).value),
        float(sheet.cell(row=i, column=4).value)
    )
    if a > 45:
        k += 1
    if a > 50:
        w += 1
    if a > 60:
        m += 1
    if a > 70:
        p += 1
    if a > 80:
        r += 1
    i += 1
print('45% -', k / (i - 2))
print('50% -', w / (i - 2))
print('60% -', m / (i - 2))
print('70% -', p / (i - 2))
print('80% -', r / (i - 2))